from datetime import datetime, timedelta
from collections import defaultdict
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from database import get_db_connection, init_db
from io import BytesIO
import openpyxl
from fpdf import FPDF
import random
import re

reset_otps = {}

# System technical integer handling capacity limit (prevents crashes/overflow)
SYSTEM_MAX_PARAMETER_LIMIT = 1000000

# Maximum catalog limit of products allowed in inventory
MAX_PRODUCT_CATALOG_LIMIT = 2500

# Per-product maximum stock capacity limit
MAX_STOCK_PER_PRODUCT = 2500
MAX_STOCK_LIMIT = MAX_STOCK_PER_PRODUCT

import os
import sys

if getattr(sys, 'frozen', False):
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    static_folder = os.path.join(sys._MEIPASS, 'static')
    app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
else:
    app = Flask(__name__)

app.secret_key = 'super_secret_key_farmacia'
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config['TEMPLATES_AUTO_RELOAD'] = True

# Ensure database is initialized
init_db()

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return '127.0.0.1'

@app.context_processor
def inject_system_info():
    return {'local_ip': get_local_ip()}

def parse_date_range(frequency, period_week, period_month, period_year):
    if frequency == 'weekly' and period_week:
        try:
            year, week = period_week.split('-W')
            # fromisocalendar is Monday start of week
            start_date = datetime.fromisocalendar(int(year), int(week), 1)
            end_date = start_date + timedelta(days=6)
            return start_date, end_date
        except Exception:
            pass
    elif frequency == 'monthly' and period_month:
        try:
            year, month = period_month.split('-')
            start_date = datetime(int(year), int(month), 1)
            if int(month) == 12:
                end_date = datetime(int(year) + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(int(year), int(month) + 1, 1) - timedelta(days=1)
            return start_date, end_date
        except Exception:
            pass
    elif frequency == 'annual' and period_year:
        try:
            start_date = datetime(int(period_year), 1, 1)
            end_date = datetime(int(period_year), 12, 31)
            return start_date, end_date
        except Exception:
            pass
    return None, None


def filter_by_date(row_date_str, start_dt, end_dt):
    try:
        dt = datetime.strptime(row_date_str, '%m/%d/%Y')
        return start_dt <= dt <= end_dt
    except Exception:
        return False

def add_notification_with_conn(conn, type, title, subtitle, color):
    now_str = datetime.now().strftime('%Y-%m-%d %I:%M %p')
    existing = conn.execute('SELECT id FROM notifications WHERE title = ?', (title,)).fetchone()
    if not existing:
        conn.execute('''
            INSERT INTO notifications (type, title, subtitle, color, created_at, seen)
            VALUES (?, ?, ?, ?, ?, 0)
        ''', (type, title, subtitle, color, now_str))

def add_notification(type, title, subtitle, color):
    conn = get_db_connection()
    add_notification_with_conn(conn, type, title, subtitle, color)
    conn.commit()
    conn.close()

def log_activity(conn, action, target_type, target_id, target_name, performed_by=None):
    """Log an action to the activity_log table."""
    try:
        by = performed_by or session.get('name', 'System')
        now_str = datetime.now().strftime('%Y-%m-%d %I:%M %p')
        conn.execute(
            'INSERT INTO activity_log (action, target_type, target_id, target_name, performed_by, performed_at) VALUES (?, ?, ?, ?, ?, ?)',
            (action, target_type, target_id, target_name, by, now_str)
        )
    except Exception as e:
        print('log_activity error:', e)

@app.context_processor
def inject_settings():
    conn = get_db_connection()
    try:
        settings_raw = conn.execute('SELECT * FROM settings').fetchall()
        settings_dict = {r['key']: r['value'] for r in settings_raw}
    except Exception:
        settings_dict = {}
        
    # Check batches for near expiry / expired and auto-insert alerts
    try:
        ref_date = datetime(2026, 3, 22)
        batches = conn.execute('''
            SELECT b.id as batch_id, i.brand as medicine, b.expiry_date, b.current_qty
            FROM batches b
            JOIN inventory i ON b.medicine_id = i.id
            WHERE b.current_qty > 0
        ''').fetchall()
        for b in batches:
            try:
                expiry_dt = datetime.strptime(b['expiry_date'], '%m/%d/%Y')
                days_left = (expiry_dt - ref_date).days
                if days_left <= 0:
                    add_notification_with_conn(conn, 'expiry', f"{b['medicine']} — Expired", f"Batch {b['batch_id']} expired on {b['expiry_date']}", '#ef4444')
                elif days_left <= 90:
                    add_notification_with_conn(conn, 'expiry', f"{b['medicine']} — Near Expiry", f"Expires in {days_left} days · Batch {b['batch_id']}", '#ffb800')
            except Exception:
                pass
        conn.commit()
    except Exception as e:
        print("Error auto-checking expiry batches:", e)
        
    # Fetch notifications from table
    notifications = []
    unseen_count = 0
    try:
        notifications_raw = conn.execute('SELECT * FROM notifications ORDER BY id DESC').fetchall()
        for r in notifications_raw:
            notifications.append({
                'id': r['id'],
                'type': r['type'],
                'title': r['title'],
                'subtitle': r['subtitle'],
                'color': r['color'],
                'created_at': r['created_at'],
                'seen': r['seen']
            })
        unseen_count = conn.execute('SELECT COUNT(*) FROM notifications WHERE seen = 0').fetchone()[0]
    except Exception as e:
        print("Error fetching notifications:", e)
        
    conn.close()
    return dict(app_settings=settings_dict, app_notifications=notifications, app_unseen_count=unseen_count)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        # Validation 1: Maximum length of 15 characters
        if len(username) > 15 or len(password) > 15:
            flash('Username and password cannot exceed 15 characters.', 'danger')
            return render_template('login.html')
            
        # Validation 2: No special characters allowed (alphanumeric only)
        if not re.match(r'^[a-zA-Z0-9]{1,15}$', username) or not re.match(r'^[a-zA-Z0-9]{1,15}$', password):
            flash('Username and password must contain only letters and numbers, maximum 15 characters (no special characters allowed).', 'danger')
            return render_template('login.html')
            
        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE username = ? AND password = ?', (username, password)).fetchone()
        conn.close()
        
        if user:
            session['user_id'] = user['id']
            session['name'] = user['name']
            session['role'] = user['role']
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid credentials', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/send-otp', methods=['POST'])
def send_otp():
    email = request.form.get('email', '').strip()
    if not email:
        return jsonify({'success': False, 'message': 'Please enter your registered email address.'})
        
    conn = get_db_connection()
    user = conn.execute('SELECT * FROM users WHERE LOWER(email) = LOWER(?) OR LOWER(username) = LOWER(?)', (email, email)).fetchone()
    conn.close()
    
    if not user:
        return jsonify({'success': False, 'message': 'No registered account found with that email address.'})
        
    otp = f"{random.randint(100000, 999999)}"
    reset_otps[user['email'].lower()] = {
        'otp': otp,
        'user_id': user['id'],
        'email': user['email']
    }
    
    print(f"\n==========================================")
    print(f"  [OTP DISPATCH] Account: {user['username']} ({user['email']})")
    print(f"  [OTP DISPATCH] Generated OTP Code: {otp}")
    print(f"==========================================\n")
    
    return jsonify({
        'success': True,
        'message': f"OTP code has been sent to {user['email']}",
        'email': user['email'],
        'otp_demo': otp
    })

@app.route('/verify-otp', methods=['POST'])
def verify_otp():
    email = request.form.get('email', '').strip().lower()
    otp_entered = request.form.get('otp', '').strip()
    
    record = reset_otps.get(email)
    if not record:
        return jsonify({'success': False, 'message': 'OTP expired or not requested. Please request a new OTP.'})
        
    if record['otp'] == otp_entered:
        session['otp_verified_user_id'] = record['user_id']
        return jsonify({'success': True, 'user_id': record['user_id']})
    else:
        return jsonify({'success': False, 'message': 'Invalid OTP code. Please check and try again.'})

@app.route('/forgot-password', methods=['POST'])
def forgot_password():
    user_id = request.form.get('user_id', '').strip()
    new_password = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()
    
    verified_session_user = session.get('otp_verified_user_id')
    if not user_id or str(user_id) != str(verified_session_user):
        flash('Security verification incomplete. Please verify your OTP code first.', 'danger')
        return redirect(url_for('login'))
        
    if confirm_password and new_password != confirm_password:
        flash('Passwords do not match. Please try again.', 'danger')
        return redirect(url_for('login'))
        
    conn = get_db_connection()
    user = conn.execute('SELECT id FROM users WHERE id = ?', (user_id,)).fetchone()
    if user:
        conn.execute('UPDATE users SET password = ? WHERE id = ?', (new_password, user['id']))
        conn.commit()
        conn.close()
        session.pop('otp_verified_user_id', None)
        flash('Password reset successfully! You can now log in with your new password.', 'success')
    else:
        conn.close()
        flash('User account not found.', 'danger')
        
    return redirect(url_for('login'))

@app.before_request
def require_login():
    allowed_routes = ['login', 'static', 'forgot_password', 'send_otp', 'verify_otp']
    if request.endpoint not in allowed_routes and 'user_id' not in session:
        return redirect(url_for('login'))

@app.route('/')
def dashboard():
    conn = get_db_connection()
    # Stats
    total_medicines = conn.execute("SELECT COUNT(*) FROM inventory WHERE archived_at IS NULL AND deleted_at IS NULL").fetchone()[0]
    low_stock = conn.execute("SELECT COUNT(*) FROM inventory WHERE status IN ('Low Stock', 'No Stock') AND archived_at IS NULL AND deleted_at IS NULL").fetchone()[0]
    from datetime import datetime
    all_batches_raw = conn.execute("SELECT expiry_date FROM batches").fetchall()
    near_expiry = 0
    for b in all_batches_raw:
        try:
            expiry_dt = datetime.strptime(b['expiry_date'], '%m/%d/%Y')
            days_left = (expiry_dt - datetime.now()).days
            if 0 < days_left <= 50:
                near_expiry += 1
        except Exception:
            pass

    sales_recorded = conn.execute('SELECT COUNT(*) FROM sales').fetchone()[0]
    po_receiving = conn.execute("SELECT COUNT(*) FROM purchase_orders WHERE status = 'For Receiving'").fetchone()[0]
    try:
        pending_expiry = conn.execute("SELECT COUNT(*) FROM batches WHERE expiry_pending = 1").fetchone()[0]
    except Exception:
        pending_expiry = 0
    
    # Tables
    low_stock_meds = conn.execute("SELECT brand as medicine, stock, reorder_point, status FROM inventory WHERE status IN ('Low Stock', 'No Stock') AND archived_at IS NULL AND deleted_at IS NULL LIMIT 5").fetchall()
    
    # For near expiry medicines table, join with inventory to get the brand name 
    near_expiry_meds_raw = conn.execute('''
        SELECT i.brand as medicine, b.expiry_date, b.status 
        FROM batches b 
        JOIN inventory i ON b.medicine_id = i.id
    ''').fetchall()
    
    near_expiry_meds = []
    for b in near_expiry_meds_raw:
        try:
            expiry_dt = datetime.strptime(b['expiry_date'], '%m/%d/%Y')
            days_left = (expiry_dt - datetime.now()).days
        except Exception:
            days_left = 71
        
        if 0 < days_left <= 50:
            near_expiry_meds.append({
                'medicine': b['medicine'],
                'expiry_date': b['expiry_date'],
                'days_left': f"{days_left} days",
                'status': 'Near Expiry'
            })
    near_expiry_meds = near_expiry_meds[:5]

    # Sales by date for Line Graph (with Last Week, Monthly, and Annual summarization)
    all_sales = conn.execute('SELECT sale_date, qty FROM sales').fetchall()
    from collections import defaultdict
    from datetime import datetime, timedelta

    sales_dates = []
    for s in all_sales:
        try:
            dt = datetime.strptime(s['sale_date'], '%m/%d/%Y')
            sales_dates.append((dt, s['qty']))
        except Exception:
            pass

    if sales_dates:
        anchor_date = max(d[0] for d in sales_dates)
    else:
        anchor_date = datetime.now()

    # 1. Last Week (last 7 days ending at anchor_date)
    last_week_data = {}
    for i in range(7):
        d = anchor_date - timedelta(days=6-i)
        last_week_data[d.strftime('%m/%d/%Y')] = 0
    
    for dt, qty in sales_dates:
        d_str = dt.strftime('%m/%d/%Y')
        if d_str in last_week_data:
            last_week_data[d_str] += qty
            
    last_week_labels = list(last_week_data.keys())
    last_week_values = list(last_week_data.values())

    # 2. This Week (daily sales from the Monday of anchor_date's week to Sunday/anchor_date)
    # anchor_date.weekday() returns 0 for Monday, 6 for Sunday
    monday_of_week = anchor_date - timedelta(days=anchor_date.weekday())
    this_week_data = {}
    for i in range(7):
        d = monday_of_week + timedelta(days=i)
        this_week_data[d.strftime('%m/%d/%Y')] = 0
        
    for dt, qty in sales_dates:
        d_str = dt.strftime('%m/%d/%Y')
        if d_str in this_week_data:
            this_week_data[d_str] += qty
            
    this_week_labels = list(this_week_data.keys())
    this_week_values = list(this_week_data.values())



    # 4. Monthly (group by Month/Year)
    monthly_map = defaultdict(int)
    for dt, qty in sales_dates:
        monthly_map[dt.strftime('%b %Y')] += qty
        
    sorted_months = sorted(monthly_map.keys(), key=lambda x: datetime.strptime(x, '%b %Y'))
    monthly_labels = sorted_months
    monthly_values = [monthly_map[m] for m in sorted_months]

    # 5. Annual (group by Year)
    annual_map = defaultdict(int)
    for dt, qty in sales_dates:
        annual_map[dt.strftime('%Y')] += qty
        
    sorted_years = sorted(annual_map.keys())
    annual_labels = sorted_years
    annual_values = [annual_map[y] for y in sorted_years]

    # Stocks for Pie Chart
    inventory_stocks = conn.execute('SELECT brand, stock FROM inventory WHERE archived_at IS NULL AND deleted_at IS NULL').fetchall()
    stock_labels = [item['brand'] for item in inventory_stocks]
    stock_values = [item['stock'] for item in inventory_stocks]

    # Recent activities
    try:
        recent_activities_raw = conn.execute(
            'SELECT * FROM activity_log ORDER BY id DESC LIMIT 8'
        ).fetchall()
        recent_activities = [dict(r) for r in recent_activities_raw]
    except Exception:
        recent_activities = []

    conn.close()

    
    return render_template('dashboard.html', 
        total_medicines=total_medicines,
        low_stock=low_stock,
        near_expiry=near_expiry,
        sales_recorded=sales_recorded,
        po_receiving=po_receiving,
        low_stock_meds=low_stock_meds,
        near_expiry_meds=near_expiry_meds,
        last_week_labels=last_week_labels,
        last_week_values=last_week_values,
        this_week_labels=this_week_labels,
        this_week_values=this_week_values,
        weekly_labels=None,
        weekly_values=None,
        monthly_labels=monthly_labels,
        monthly_values=monthly_values,
        annual_labels=annual_labels,
        annual_values=annual_values,
        stock_labels=stock_labels,
        stock_values=stock_values,
        recent_activities=recent_activities,
        pending_expiry=pending_expiry,
        active_page='dashboard'
    )

@app.route('/inventory', methods=['GET', 'POST'])
def inventory():
    conn = get_db_connection()
    if request.method == 'POST':
        prod_count = conn.execute("SELECT COUNT(*) FROM inventory WHERE archived_at IS NULL AND deleted_at IS NULL").fetchone()[0]
        if prod_count >= MAX_PRODUCT_CATALOG_LIMIT:
            conn.close()
            flash('Maximum product limit reached (2,500 products max). Cannot add additional products.', 'error')
            return redirect(url_for('inventory'))

        generic = request.form.get('generic', '')
        brand = request.form['brand']
        category = request.form['category']
        reorder_point = min(MAX_STOCK_LIMIT, max(0, int(request.form.get('reorder_point', 0))))
        
        product_name = request.form.get('product_name', brand)
        description = request.form.get('description', '')
        unit_of_measure = request.form.get('unit_of_measure', 'Piece')
        purchase_price = float(request.form.get('purchase_price', 0.0))
        supplier = request.form.get('supplier', '')
        
        # Generate next medicine ID (e.g. MED-034)
        last_med = conn.execute("SELECT id FROM inventory ORDER BY id DESC LIMIT 1").fetchone()
        if last_med:
            try:
                last_num = int(last_med['id'].split('-')[1])
                next_med_id = f"MED-{last_num + 1:03d}"
            except (IndexError, ValueError):
                next_med_id = "MED-034"
        else:
            next_med_id = "MED-001"
            
        conn.execute("""
            INSERT INTO inventory (id, generic, brand, category, stock, reorder_point, status, product_name, description, unit_of_measure, purchase_price, supplier) 
            VALUES (?, ?, ?, ?, 0, ?, 'No Stock', ?, ?, ?, ?, ?)
        """, (next_med_id, generic, brand, category, reorder_point, product_name, description, unit_of_measure, purchase_price, supplier))
        log_activity(conn, 'add', 'product', next_med_id, brand)
        conn.commit()
        conn.close()
        return redirect(url_for('inventory'))
        
    items = conn.execute('SELECT * FROM inventory WHERE archived_at IS NULL AND deleted_at IS NULL').fetchall()
    suppliers = conn.execute('SELECT name FROM suppliers').fetchall()
    conn.close()
    return render_template('inventory.html', items=items, suppliers=suppliers, active_page='inventory')

@app.route('/inventory/edit', methods=['POST'])
def edit_medicine():
    conn = get_db_connection()
    med_id = request.form['id']
    generic = request.form.get('generic', '')
    brand = request.form['brand']
    category = request.form['category']
    reorder_point = min(MAX_STOCK_LIMIT, max(0, int(request.form.get('reorder_point', 0))))
    
    product_name = request.form.get('product_name', brand)
    description = request.form.get('description', '')
    unit_of_measure = request.form.get('unit_of_measure', 'Piece')
    purchase_price = float(request.form.get('purchase_price', 0.0))
    supplier = request.form.get('supplier', '')
    
    # Recalculate status
    med = conn.execute("SELECT stock FROM inventory WHERE id = ?", (med_id,)).fetchone()
    stock = med['stock'] if med else 0
    new_status = 'No Stock' if stock == 0 else ('Low Stock' if stock < reorder_point else 'Good')
    
    conn.execute('''
        UPDATE inventory 
        SET generic = ?, brand = ?, category = ?, reorder_point = ?, status = ?,
            product_name = ?, description = ?, unit_of_measure = ?, purchase_price = ?, supplier = ?
        WHERE id = ?
    ''', (generic, brand, category, reorder_point, new_status, product_name, description, unit_of_measure, purchase_price, supplier, med_id))
    log_activity(conn, 'edit', 'product', med_id, brand)
    conn.commit()
    conn.close()
    flash('Medicine updated successfully!')
    return redirect(url_for('inventory'))

# ─── ARCHIVE / TRASH / RESTORE ROUTES ───────────────────────────────────────

@app.route('/inventory/archive/<med_id>', methods=['POST'])
def archive_medicine(med_id):
    if session.get('role') not in ['Owner / Pharmacist']:
        flash('Only owners can archive products.')
        return redirect(url_for('inventory'))
    conn = get_db_connection()
    med = conn.execute('SELECT brand FROM inventory WHERE id = ?', (med_id,)).fetchone()
    if med:
        now_str = datetime.now().strftime('%Y-%m-%d %I:%M %p')
        conn.execute('UPDATE inventory SET archived_at = ?, deleted_at = NULL WHERE id = ?', (now_str, med_id))
        log_activity(conn, 'archive', 'product', med_id, med['brand'])
        conn.commit()
    conn.close()
    flash(f'{med["brand"]} has been archived.')
    return redirect(url_for('inventory'))

@app.route('/inventory/restore/<med_id>', methods=['POST'])
def restore_medicine(med_id):
    if session.get('role') not in ['Owner / Pharmacist']:
        flash('Only owners can restore products.')
        return redirect(url_for('history'))
    conn = get_db_connection()
    med = conn.execute('SELECT brand FROM inventory WHERE id = ?', (med_id,)).fetchone()
    if med:
        conn.execute('UPDATE inventory SET archived_at = NULL, deleted_at = NULL WHERE id = ?', (med_id,))
        log_activity(conn, 'restore', 'product', med_id, med['brand'])
        conn.commit()
    conn.close()
    flash(f'{med["brand"]} has been restored to active inventory.')
    return redirect(url_for('history'))

@app.route('/inventory/trash/<med_id>', methods=['POST'])
def trash_medicine(med_id):
    if session.get('role') not in ['Owner / Pharmacist']:
        flash('Only owners can delete products.')
        return redirect(url_for('history'))
    conn = get_db_connection()
    med = conn.execute('SELECT brand FROM inventory WHERE id = ?', (med_id,)).fetchone()
    if med:
        now_str = datetime.now().strftime('%Y-%m-%d %I:%M %p')
        conn.execute('UPDATE inventory SET deleted_at = ?, archived_at = NULL WHERE id = ?', (now_str, med_id))
        log_activity(conn, 'trash', 'product', med_id, med['brand'])
        conn.commit()
    conn.close()
    flash(f'{med["brand"]} moved to trash. It will be permanently deleted after 30 days.')
    return redirect(url_for('history'))

@app.route('/inventory/delete-permanent/<med_id>', methods=['POST'])
def delete_medicine_permanent(med_id):
    if session.get('role') not in ['Owner / Pharmacist']:
        flash('Only owners can permanently delete products.')
        return redirect(url_for('history'))
    conn = get_db_connection()
    med = conn.execute('SELECT brand FROM inventory WHERE id = ?', (med_id,)).fetchone()
    if med:
        log_activity(conn, 'delete', 'product', med_id, med['brand'])
        conn.execute('DELETE FROM inventory WHERE id = ?', (med_id,))
        conn.commit()
    conn.close()
    flash(f'{med["brand"]} has been permanently deleted.')
    return redirect(url_for('history'))

@app.route('/history')
def history():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db_connection()
    now = datetime.now()

    # Auto-purge products whose deleted_at > 30 days
    try:
        all_trashed = conn.execute("SELECT id, brand, deleted_at FROM inventory WHERE deleted_at IS NOT NULL").fetchall()
        for t in all_trashed:
            try:
                deleted_dt = datetime.strptime(t['deleted_at'], '%Y-%m-%d %I:%M %p')
                if (now - deleted_dt).days >= 30:
                    log_activity(conn, 'delete', 'product', t['id'], t['brand'], performed_by='System (Auto-Purge)')
                    conn.execute('DELETE FROM inventory WHERE id = ?', (t['id'],))
            except Exception:
                pass
        conn.commit()
    except Exception as e:
        print('Auto-purge error:', e)

    # Archived products
    archived = conn.execute(
        "SELECT * FROM inventory WHERE archived_at IS NOT NULL AND deleted_at IS NULL ORDER BY archived_at DESC"
    ).fetchall()

    # Trashed products with days_remaining
    trashed_raw = conn.execute(
        "SELECT * FROM inventory WHERE deleted_at IS NOT NULL ORDER BY deleted_at DESC"
    ).fetchall()
    trashed = []
    for t in trashed_raw:
        try:
            deleted_dt = datetime.strptime(t['deleted_at'], '%Y-%m-%d %I:%M %p')
            days_left = 30 - (now - deleted_dt).days
            days_left = max(0, days_left)
        except Exception:
            days_left = 30
        trashed.append({'item': dict(t), 'days_remaining': days_left})

    # Activity log (paginated — last 100)
    activity_log_raw = conn.execute(
        'SELECT * FROM activity_log ORDER BY id DESC LIMIT 100'
    ).fetchall()
    activity_log = [dict(r) for r in activity_log_raw]

    conn.close()
    return render_template('history.html',
        archived=archived,
        trashed=trashed,
        activity_log=activity_log,
        active_page='history'
    )


@app.route('/batches')
def batches():
    conn = get_db_connection()
    items = conn.execute('''
        SELECT b.id as batch_id, i.brand as medicine, b.expiry_date, b.current_qty, b.status, i.category
        FROM batches b 
        JOIN inventory i ON b.medicine_id = i.id
    ''').fetchall()
    conn.close()
    
    list_items = []
    for b in items:
        try:
            expiry_dt = datetime.strptime(b['expiry_date'], '%m/%d/%Y')
            days_left = (expiry_dt - datetime.now()).days
            if days_left <= 0:
                status = 'Expired'
            elif days_left <= 50:
                status = 'Near Expiry'
            else:
                status = 'Good'
        except Exception:
            status = 'Good'
            
        list_items.append({
            'batch_id': b['batch_id'],
            'medicine': b['medicine'],
            'expiry_date': b['expiry_date'],
            'current_qty': b['current_qty'],
            'status': status,
            'category': b['category']
        })
    return render_template('batches.html', items=list_items, active_page='batches')

@app.route('/suppliers', methods=['GET', 'POST'])
def suppliers():
    conn = get_db_connection()
    if request.method == 'POST':
        name = request.form['name']
        address = request.form['address']
        contact = request.form['contact']
        
        # Generate next Supplier ID (e.g. SUP-003)
        last_sup = conn.execute("SELECT id FROM suppliers ORDER BY id DESC LIMIT 1").fetchone()
        if last_sup:
            try:
                last_num = int(last_sup['id'].split('-')[1])
                next_sup_id = f"SUP-{last_num + 1:03d}"
            except (IndexError, ValueError):
                next_sup_id = "SUP-001"
        else:
            next_sup_id = "SUP-001"
            
        conn.execute("INSERT INTO suppliers (id, name, address, contact) VALUES (?, ?, ?, ?)",
                     (next_sup_id, name, address, contact))
        conn.commit()
        conn.close()
        return redirect(url_for('suppliers'))
        
    items = conn.execute('SELECT * FROM suppliers').fetchall()
    conn.close()
    return render_template('suppliers.html', items=items, active_page='suppliers')

@app.route('/suppliers/edit', methods=['POST'])
def edit_supplier():
    if session.get('role') not in ['Owner / Pharmacist']:
        flash('Only owners can edit suppliers.')
        return redirect(url_for('suppliers'))
    conn = get_db_connection()
    sup_id = request.form['id']
    name = request.form['name']
    address = request.form['address']
    contact = request.form['contact']
    
    conn.execute('''
        UPDATE suppliers 
        SET name = ?, address = ?, contact = ?
        WHERE id = ?
    ''', (name, address, contact, sup_id))
    conn.commit()
    conn.close()
    flash('Supplier updated successfully!')
    return redirect(url_for('suppliers'))

@app.route('/suppliers/<sup_id>/delete', methods=['POST'])
def delete_supplier(sup_id):
    if session.get('role') not in ['Owner / Pharmacist']:
        flash('Only owners can delete suppliers.')
        return redirect(url_for('suppliers'))
    conn = get_db_connection()
    po = conn.execute("SELECT * FROM purchase_orders WHERE supplier_id = ?", (sup_id,)).fetchone()
    if po:
        flash("Cannot delete supplier: linked purchase orders exist!")
    else:
        conn.execute("DELETE FROM suppliers WHERE id = ?", (sup_id,))
        conn.commit()
        flash("Supplier deleted successfully!")
    conn.close()
    return redirect(url_for('suppliers'))

@app.route('/purchase_orders', methods=['GET', 'POST'])
def purchase_orders():
    conn = get_db_connection()
    
    if request.method == 'POST':
        supplier_id = request.form['supplier_id']
        notes = request.form['notes']
        medicines = request.form.getlist('medicine[]')
        quantities = request.form.getlist('quantity[]')
        
        # Generate next PO ID
        last_po = conn.execute("SELECT id FROM purchase_orders ORDER BY id DESC LIMIT 1").fetchone()
        if last_po:
            try:
                last_num = int(last_po['id'].split('-')[1])
                next_po_id = f"PO-{last_num + 1:03d}"
            except (IndexError, ValueError):
                next_po_id = "PO-001"
        else:
            next_po_id = "PO-001"
            
        order_date = datetime.now().strftime('%m/%d/%Y')
        prepared_by = session.get('name', 'Owner')
        
        # Insert PO
        conn.execute('''
            INSERT INTO purchase_orders (id, supplier_id, prepared_by, order_date, status, notes) 
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (next_po_id, supplier_id, prepared_by, order_date, 'For Receiving', notes))
        
        # Insert PO Items
        for med_id, qty in zip(medicines, quantities):
            if med_id and qty:
                sanitized_qty = min(MAX_STOCK_LIMIT, max(1, int(qty)))
                conn.execute('''
                    INSERT INTO purchase_order_items (purchase_order_id, medicine_id, quantity) 
                    VALUES (?, ?, ?)
                ''', (next_po_id, med_id, sanitized_qty))
                
        # Add PO pending notification
        supplier_name = conn.execute("SELECT name FROM suppliers WHERE id = ?", (supplier_id,)).fetchone()['name']
        valid_items = [m for m in medicines if m]
        add_notification_with_conn(conn, 'po', f"PO {next_po_id} pending receipt", f"From {supplier_name} · {len(valid_items)} items · For Receiving", '#3b82f6')
        log_activity(conn, 'add', 'purchase_order', next_po_id, f"PO {next_po_id} ({supplier_name})")
        
        conn.commit()
        conn.close()
        return redirect(url_for('purchase_orders'))
        
    # GET
    suppliers = conn.execute('SELECT * FROM suppliers').fetchall()
    medicines = conn.execute('SELECT * FROM inventory').fetchall()
    
    # Fetch POs with individual items
    orders_raw = conn.execute('''
        SELECT po.id, po.supplier_id, s.name as supplier, po.prepared_by, po.order_date, po.status, po.notes, po.received_by,
               i.brand as medicine_name, poi.quantity as item_qty
        FROM purchase_order_items poi
        JOIN purchase_orders po ON poi.purchase_order_id = po.id
        JOIN suppliers s ON po.supplier_id = s.id
        JOIN inventory i ON poi.medicine_id = i.id
        ORDER BY po.id DESC, poi.id ASC
    ''').fetchall()
    
    orders = []
    for o in orders_raw:
        notes_escaped = (o['notes'] or '').replace("'", "\\'").replace("\n", " ")
        orders.append({
            'id': o['id'],
            'supplier_id': o['supplier_id'],
            'supplier': o['supplier'],
            'prepared_by': o['prepared_by'],
            'order_date': o['order_date'],
            'status': o['status'],
            'notes': notes_escaped,
            'products_list': o['medicine_name'],
            'total_quantity': o['item_qty'],
            'received_by': o['received_by'] or ''
        })
    
    # Fetch detailed PO items for the Details Modal
    po_items_raw = conn.execute('''
        SELECT poi.purchase_order_id, poi.medicine_id, i.brand as medicine_name, i.generic as generic_name, poi.quantity
        FROM purchase_order_items poi
        JOIN inventory i ON poi.medicine_id = i.id
    ''').fetchall()
    
    po_items_map = {}
    for item in po_items_raw:
        po_id = item['purchase_order_id']
        if po_id not in po_items_map:
            po_items_map[po_id] = []
        po_items_map[po_id].append({
            'medicine_id': item['medicine_id'],
            'medicine_name': item['medicine_name'],
            'generic_name': item['generic_name'],
            'quantity': item['quantity']
        })
        
    conn.close()
    return render_template('purchase_orders.html', 
                           suppliers=suppliers, 
                           medicines=medicines, 
                           orders=orders, 
                           po_items_map=po_items_map,
                           active_page='purchase_orders')

@app.route('/purchase_orders/<po_id>/items-json')
def po_items_json(po_id):
    """Return PO items as JSON for the expiry date modal."""
    if 'user_id' not in session:
        from flask import jsonify
        return jsonify({'items': []})
    conn = get_db_connection()
    items = conn.execute('''
        SELECT poi.medicine_id, i.brand as medicine_name, poi.quantity
        FROM purchase_order_items poi
        JOIN inventory i ON poi.medicine_id = i.id
        WHERE poi.purchase_order_id = ?
    ''', (po_id,)).fetchall()
    conn.close()
    from flask import jsonify
    return jsonify({'items': [dict(r) for r in items]})

@app.route('/purchase_orders/<po_id>/receive', methods=['POST'])
def receive_po(po_id):
    conn = get_db_connection()
    # Ensure expiry_pending column exists
    try:
        conn.execute("ALTER TABLE batches ADD COLUMN expiry_pending INTEGER DEFAULT 0")
        conn.commit()
    except Exception:
        pass  # already exists

    po = conn.execute("SELECT * FROM purchase_orders WHERE id = ?", (po_id,)).fetchone()
    if not po or po['status'] != 'For Receiving':
        conn.close()
        return redirect(url_for('purchase_orders'))

    receiver = session.get('name', 'Unknown')
    skipped  = request.form.get('skip_expiry') == '1'

    conn.execute("UPDATE purchase_orders SET status = 'Received', received_by = ? WHERE id = ?", (receiver, po_id))
    items = conn.execute("SELECT * FROM purchase_order_items WHERE purchase_order_id = ?", (po_id,)).fetchall()

    for item in items:
        med_id = item['medicine_id']
        qty    = min(MAX_STOCK_LIMIT, max(0, int(item['quantity'])))

        med = conn.execute("SELECT stock, reorder_point, brand FROM inventory WHERE id = ?", (med_id,)).fetchone()
        current_stock = med['stock'] if med else 0
        new_stock = min(MAX_STOCK_LIMIT, current_stock + qty)
        conn.execute("UPDATE inventory SET stock = ? WHERE id = ?", (new_stock, med_id))
        if med:
            new_status = 'No Stock' if new_stock == 0 else ('Low Stock' if new_stock < med['reorder_point'] else 'Good')
            conn.execute("UPDATE inventory SET status = ? WHERE id = ?", (new_status, med_id))
            if new_stock >= 2500:
                add_notification_with_conn(conn, 'overstock', f"⚠️ Overstock Warning: {med['brand']}", f"Stock level reached {new_stock} units. Storage capacity threshold (2,500 units) exceeded!", '#f59e0b')
                flash(f"⚠️ Storage Capacity Warning: {med['brand']} stock is now {new_stock} units, exceeding storage capacity threshold (2,500 units)!", "warning")

        last_batch = conn.execute("SELECT id FROM batches ORDER BY id DESC LIMIT 1").fetchone()
        if last_batch:
            try:
                last_num = int(last_batch['id'].split('-')[1])
                next_batch_id = f"BAT-{last_num + 1:03d}"
            except (IndexError, ValueError):
                next_batch_id = "BAT-004"
        else:
            next_batch_id = "BAT-001"

        if skipped:
            expiry_date    = ''   # blank — user must fill in later
            expiry_pending = 1
        else:
            raw = request.form.get(f'expiry_{med_id}', '').strip()
            if raw:
                try:
                    dt = datetime.strptime(raw, '%Y-%m-%d')
                    expiry_date = dt.strftime('%m/%d/%Y')
                except Exception:
                    expiry_date = raw
                expiry_pending = 0
            else:
                expiry_date = ''
                expiry_pending = 1

        conn.execute(
            "INSERT INTO batches (id, medicine_id, expiry_date, current_qty, status, expiry_pending) VALUES (?, ?, ?, ?, ?, ?)",
            (next_batch_id, med_id, expiry_date, qty, 'Good', expiry_pending)
        )

        last_movement = conn.execute("SELECT id FROM stock_movements ORDER BY id DESC LIMIT 1").fetchone()
        if last_movement:
            try:
                last_num = int(last_movement['id'].split('-')[1])
                next_trn_id = f"TRN-{last_num + 1:03d}"
            except (IndexError, ValueError):
                next_trn_id = "TRN-002"
        else:
            next_trn_id = "TRN-001"

        trn_date = datetime.now().strftime('%Y-%m-%d %I:%M %p')
        conn.execute(
            "INSERT INTO stock_movements (id, type, medicine_id, batch_id, quantity, movement_date, reference) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (next_trn_id, 'Stock-In', med_id, next_batch_id, qty, trn_date, f'PO {po_id}')
        )

    add_notification_with_conn(conn, 'po', f"PO {po_id} received", "Stock levels have been successfully updated", '#10b981')
    log_activity(conn, 'po_received', 'purchase_order', po_id, f"PO {po_id}")
    if skipped:
        add_notification_with_conn(conn, 'warning', f"Expiry dates pending for PO {po_id}", "Please set expiry dates in Expiry Monitoring", '#f59e0b')
    conn.commit()
    conn.close()
    return redirect(url_for('purchase_orders'))

@app.route('/batches/<batch_id>/set-expiry', methods=['POST'])
def set_batch_expiry(batch_id):
    """Update expiry date on a pending batch from Expiry Monitoring page."""
    conn = get_db_connection()
    expiry_raw = request.form.get('expiry_date', '').strip()
    if expiry_raw:
        try:
            dt = datetime.strptime(expiry_raw, '%Y-%m-%d')
            expiry_date = dt.strftime('%m/%d/%Y')
        except Exception:
            expiry_date = expiry_raw
        conn.execute("UPDATE batches SET expiry_date = ?, expiry_pending = 0 WHERE id = ?", (expiry_date, batch_id))
        conn.commit()
        flash(f'Expiry date updated for batch {batch_id}.')
    conn.close()
    return redirect(url_for('expiry_monitoring'))


@app.route('/purchase_orders/edit', methods=['POST'])
def edit_po():
    conn = get_db_connection()
    po_id = request.form['id']
    supplier_id = request.form['supplier_id']
    notes = request.form['notes']
    medicines = request.form.getlist('medicine[]')
    quantities = request.form.getlist('quantity[]')
    
    # Delete old items
    conn.execute("DELETE FROM purchase_order_items WHERE purchase_order_id = ?", (po_id,))
    
    # Update PO
    conn.execute('''
        UPDATE purchase_orders 
        SET supplier_id = ?, notes = ?
        WHERE id = ?
    ''', (supplier_id, notes, po_id))
    
    # Insert new items
    for med_id, qty in zip(medicines, quantities):
        if med_id and qty:
            sanitized_qty = min(MAX_STOCK_LIMIT, max(1, int(qty)))
            conn.execute('''
                INSERT INTO purchase_order_items (purchase_order_id, medicine_id, quantity)
                VALUES (?, ?, ?)
            ''', (po_id, med_id, sanitized_qty))
            
    conn.commit()
    conn.close()
    flash('Purchase Order updated successfully!')
    return redirect(url_for('purchase_orders'))

@app.route('/purchase_orders/<po_id>/delete', methods=['POST'])
def delete_po(po_id):
    conn = get_db_connection()
    po = conn.execute("SELECT * FROM purchase_orders WHERE id = ?", (po_id,)).fetchone()
    if po and po['status'] == 'For Receiving':
        conn.execute("DELETE FROM purchase_order_items WHERE purchase_order_id = ?", (po_id,))
        conn.execute("DELETE FROM purchase_orders WHERE id = ?", (po_id,))
        conn.commit()
        flash('Purchase Order deleted successfully!')
    conn.close()
    return redirect(url_for('purchase_orders'))

@app.route('/sales', methods=['GET', 'POST'])
def sales():
    conn = get_db_connection()
    if request.method == 'POST':
        medicine_id = request.form['medicine_id']
        qty = min(MAX_STOCK_LIMIT, max(1, int(request.form.get('qty', 1))))
        
        # Validate stock
        med = conn.execute("SELECT * FROM inventory WHERE id = ?", (medicine_id,)).fetchone()
        if med and med['stock'] >= qty:
            remaining_to_sell = qty
            batches = conn.execute("SELECT * FROM batches WHERE medicine_id = ? AND current_qty > 0", (medicine_id,)).fetchall()
            
            # Sort in Python by parsing expiry date
            def parse_date(b):
                try:
                    return datetime.strptime(b['expiry_date'], '%m/%d/%Y')
                except Exception:
                    return datetime.max
            batches = sorted(batches, key=parse_date)
            
            for b in batches:
                if remaining_to_sell <= 0:
                    break
                sell_qty = min(remaining_to_sell, b['current_qty'])
                
                # Update batch
                conn.execute("UPDATE batches SET current_qty = current_qty - ? WHERE id = ?", (sell_qty, b['id']))
                # Update inventory stock
                conn.execute("UPDATE inventory SET stock = stock - ? WHERE id = ?", (sell_qty, medicine_id))
                
                # Recalculate status
                med_item = conn.execute("SELECT stock, reorder_point FROM inventory WHERE id = ?", (medicine_id,)).fetchone()
                if med_item:
                    new_status = 'No Stock' if med_item['stock'] == 0 else ('Low Stock' if med_item['stock'] < med_item['reorder_point'] else 'Good')
                    conn.execute("UPDATE inventory SET status = ? WHERE id = ?", (new_status, medicine_id))
                
                # Log stock movement
                last_movement = conn.execute("SELECT id FROM stock_movements ORDER BY id DESC LIMIT 1").fetchone()
                if last_movement:
                    try:
                        last_num = int(last_movement['id'].split('-')[1])
                        next_trn_id = f"TRN-{last_num + 1:03d}"
                    except (IndexError, ValueError):
                        next_trn_id = "TRN-002"
                else:
                    next_trn_id = "TRN-001"
                
                trn_date = datetime.now().strftime('%Y-%m-%d %I:%M %p')
                
                # Next SAL ID for reference
                last_sale = conn.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1").fetchone()
                if last_sale:
                    try:
                        last_num = int(last_sale['id'].split('-')[1])
                        next_sale_id = f"SAL-{last_num + 1:03d}"
                    except (IndexError, ValueError):
                        next_sale_id = "SAL-002"
                else:
                    next_sale_id = "SAL-001"
                
                conn.execute("INSERT INTO stock_movements (id, type, medicine_id, batch_id, quantity, movement_date, reference) VALUES (?, ?, ?, ?, ?, ?, ?)",
                             (next_trn_id, 'Stock-Out', medicine_id, b['id'], -sell_qty, trn_date, f'Sale {next_sale_id}'))
                
                remaining_to_sell -= sell_qty
            
            # Insert Sale
            last_sale = conn.execute("SELECT id FROM sales ORDER BY id DESC LIMIT 1").fetchone()
            if last_sale:
                try:
                    last_num = int(last_sale['id'].split('-')[1])
                    next_sale_id = f"SAL-{last_num + 1:03d}"
                except (IndexError, ValueError):
                    next_sale_id = "SAL-002"
            else:
                next_sale_id = "SAL-001"
            
            sale_date = datetime.now().strftime('%m/%d/%Y')
            sold_by = session.get('name', 'Assistant Pharmacist')
            conn.execute("INSERT INTO sales (id, medicine_id, sale_date, qty, sold_by) VALUES (?, ?, ?, ?, ?)",
                         (next_sale_id, medicine_id, sale_date, qty, sold_by))
            log_activity(conn, 'sale', 'product', medicine_id, med['brand'])
            
            # Check updated stock levels for alerts
            updated_med = conn.execute("SELECT brand, stock, reorder_point FROM inventory WHERE id = ?", (medicine_id,)).fetchone()
            if updated_med:
                if updated_med['stock'] == 0:
                    add_notification_with_conn(conn, 'low_stock', f"{updated_med['brand']} is Out of Stock", "Remaining stock is 0 units", '#ef4444')
                elif updated_med['stock'] < updated_med['reorder_point']:
                    add_notification_with_conn(conn, 'low_stock', f"{updated_med['brand']} is Low Stock", f"Only {updated_med['stock']} units left · Reorder point: {updated_med['reorder_point']}", '#ef4444')
            
            conn.commit()
        else:
            flash("Insufficient stock or medicine not found!")
        
        conn.close()
        return redirect(url_for('sales'))
        
    medicines = conn.execute("SELECT * FROM inventory").fetchall()
    sales_history = conn.execute('''
        SELECT s.id, s.medicine_id, s.sale_date, i.brand as medicine_name, s.qty, s.sold_by
        FROM sales s
        JOIN inventory i ON s.medicine_id = i.id
        ORDER BY s.id DESC
    ''').fetchall()
    conn.close()
    return render_template('sales.html', medicines=medicines, sales_history=sales_history, active_page='sales')

@app.route('/sales/edit', methods=['POST'])
def edit_sale():
    conn = get_db_connection()
    sale_id = request.form['id']
    medicine_id = request.form['medicine_id']
    new_qty = int(request.form['qty'])
    
    # Fetch old sale details
    old_sale = conn.execute("SELECT * FROM sales WHERE id = ?", (sale_id,)).fetchone()
    if not old_sale:
        conn.close()
        flash("Sale record not found!")
        return redirect(url_for('sales'))
        
    old_qty = old_sale['qty']
    old_med_id = old_sale['medicine_id']
    
    # If medicine hasn't changed:
    if old_med_id == medicine_id:
        # Check if we have enough stock to cover the difference
        med = conn.execute("SELECT stock, reorder_point, brand FROM inventory WHERE id = ?", (medicine_id,)).fetchone()
        diff = new_qty - old_qty
        if med and med['stock'] >= diff:
            # Update stock
            new_stock = med['stock'] - diff
            new_status = 'No Stock' if new_stock == 0 else ('Low Stock' if new_stock < med['reorder_point'] else 'Good')
            conn.execute("UPDATE inventory SET stock = ?, status = ? WHERE id = ?", (new_stock, new_status, medicine_id))
            # Update sale
            conn.execute("UPDATE sales SET qty = ? WHERE id = ?", (new_qty, sale_id))
            conn.commit()
            flash("Sale record updated successfully!")
        else:
            flash("Insufficient stock for the requested quantity change!")
    else:
        # Medicine changed!
        # Return old stock to old medicine
        old_med = conn.execute("SELECT stock, reorder_point FROM inventory WHERE id = ?", (old_med_id,)).fetchone()
        if old_med:
            old_new_stock = old_med['stock'] + old_qty
            old_new_status = 'No Stock' if old_new_stock == 0 else ('Low Stock' if old_new_stock < old_med['reorder_point'] else 'Good')
            conn.execute("UPDATE inventory SET stock = ?, status = ? WHERE id = ?", (old_new_stock, old_new_status, old_med_id))
            
        # Deduct new stock from new medicine
        new_med = conn.execute("SELECT stock, reorder_point FROM inventory WHERE id = ?", (medicine_id,)).fetchone()
        if new_med and new_med['stock'] >= new_qty:
            new_new_stock = new_med['stock'] - new_qty
            new_new_status = 'No Stock' if new_new_stock == 0 else ('Low Stock' if new_new_stock < new_med['reorder_point'] else 'Good')
            conn.execute("UPDATE inventory SET stock = ?, status = ? WHERE id = ?", (new_new_stock, new_new_status, medicine_id))
            
            # Update sale
            conn.execute("UPDATE sales SET medicine_id = ?, qty = ? WHERE id = ?", (medicine_id, new_qty, sale_id))
            conn.commit()
            flash("Sale record updated successfully!")
        else:
            # Rollback old medicine stock return
            conn.rollback()
            flash("Insufficient stock for the new medicine!")
            
    conn.close()
    return redirect(url_for('sales'))

@app.route('/stock_movements')
def stock_movements():
    conn = get_db_connection()
    items = conn.execute('''
        SELECT sm.id, sm.type, i.brand as medicine, sm.batch_id as batch, sm.quantity, sm.movement_date as date, sm.reference, i.category
        FROM stock_movements sm
        JOIN inventory i ON sm.medicine_id = i.id
        ORDER BY sm.id DESC
    ''').fetchall()
    conn.close()
    return render_template('stock_movements.html', items=items, active_page='stock_movements')

@app.route('/expiry_monitoring')
def expiry_monitoring():
    conn = get_db_connection()

    # Migrate: ensure disposals table exists without full reset
    conn.execute('''CREATE TABLE IF NOT EXISTS disposals (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id TEXT NOT NULL, medicine_id TEXT NOT NULL,
        medicine_name TEXT NOT NULL, qty_disposed INTEGER NOT NULL,
        reason TEXT DEFAULT 'Expired', disposed_by TEXT NOT NULL,
        disposed_at TEXT NOT NULL, notes TEXT DEFAULT ''
    )''')

    # Ensure expiry_pending column exists
    try:
        conn.execute("ALTER TABLE batches ADD COLUMN expiry_pending INTEGER DEFAULT 0")
        conn.commit()
    except Exception:
        pass

    batches = conn.execute('''
        SELECT b.id as batch_id, i.id as medicine_id, i.brand as medicine,
               b.expiry_date, b.current_qty, b.status,
               COALESCE(b.expiry_pending, 0) as expiry_pending
        FROM batches b 
        JOIN inventory i ON b.medicine_id = i.id
    ''').fetchall()

    # Fetch all disposed batch IDs for quick lookup
    disposed_ids = set(
        r['batch_id'] for r in conn.execute('SELECT batch_id FROM disposals').fetchall()
    )

    # Fetch disposal log (latest 50)
    disposal_log = conn.execute('''
        SELECT * FROM disposals ORDER BY id DESC LIMIT 50
    ''').fetchall()

    conn.close()

    list_batches = []
    for b in batches:
        is_pending = bool(b['expiry_pending']) or not b['expiry_date'] or b['expiry_date'].strip() == ''

        if is_pending:
            days_left = None
            status    = 'Pending'
        else:
            try:
                expiry_dt = datetime.strptime(b['expiry_date'], '%m/%d/%Y')
                days_left = (expiry_dt - datetime.now()).days
                if days_left <= 0:
                    days_left = 0
                    status = 'Expired'
                elif days_left <= 50:
                    status = 'Near Expiry'
                else:
                    status = 'Good'
            except Exception:
                days_left = None
                status    = 'Pending'

        list_batches.append({
            'medicine':       b['medicine'],
            'medicine_id':    b['medicine_id'],
            'batch_id':       b['batch_id'],
            'expiry_date':    b['expiry_date'] if b['expiry_date'] else '—',
            'days_left':      f"{days_left} days" if days_left is not None else '—',
            'current_qty':    b['current_qty'],
            'status':         status,
            'disposed':       b['batch_id'] in disposed_ids,
            'expiry_pending': is_pending
        })

    return render_template('expiry_monitoring.html',
        items=list_batches,
        disposal_log=[dict(r) for r in disposal_log],
        active_page='expiry_monitoring'
    )

@app.route('/expiry/dispose/<batch_id>', methods=['POST'])
def dispose_batch(batch_id):
    conn = get_db_connection()
    notes = request.form.get('notes', '').strip()

    batch = conn.execute('''
        SELECT b.id, b.current_qty, b.medicine_id, i.brand
        FROM batches b JOIN inventory i ON b.medicine_id = i.id
        WHERE b.id = ?
    ''', (batch_id,)).fetchone()

    if not batch:
        conn.close()
        flash('Batch not found.')
        return redirect(url_for('expiry_monitoring'))

    qty = batch['current_qty']
    now_str = datetime.now().strftime('%Y-%m-%d %I:%M %p')
    by = session.get('name', 'System')

    # Record disposal
    conn.execute('''
        INSERT INTO disposals (batch_id, medicine_id, medicine_name, qty_disposed, reason, disposed_by, disposed_at, notes)
        VALUES (?, ?, ?, ?, 'Expired', ?, ?, ?)
    ''', (batch_id, batch['medicine_id'], batch['brand'], qty, by, now_str, notes))

    # Zero out batch quantity
    conn.execute('UPDATE batches SET current_qty = 0, status = ? WHERE id = ?', ('Disposed', batch_id))

    # Deduct from inventory stock
    med = conn.execute('SELECT stock, reorder_point FROM inventory WHERE id = ?', (batch['medicine_id'],)).fetchone()
    if med:
        new_stock = max(0, med['stock'] - qty)
        new_status = 'No Stock' if new_stock == 0 else ('Low Stock' if new_stock < med['reorder_point'] else 'Good')
        conn.execute('UPDATE inventory SET stock = ?, status = ? WHERE id = ?', (new_stock, new_status, batch['medicine_id']))

    # Log to stock movements
    trn_id = f"DIS-{batch_id}"
    conn.execute('''
        INSERT OR IGNORE INTO stock_movements (id, type, medicine_id, batch_id, quantity, movement_date, reference)
        VALUES (?, 'Disposal', ?, ?, ?, ?, ?)
    ''', (trn_id, batch['medicine_id'], batch_id, -qty, now_str, f'Disposal of expired batch {batch_id}'))

    # Activity log
    log_activity(conn, 'dispose', 'batch', batch_id, batch['brand'])

    conn.commit()
    conn.close()
    flash(f"Batch {batch_id} ({batch['brand']}) — {qty} units disposed and recorded.")
    return redirect(url_for('expiry_monitoring'))


@app.route('/users', methods=['GET', 'POST'])
def users():
    if session.get('role') == 'Staff':
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    if request.method == 'POST':
        user_db_id = request.form.get('id')
        name = request.form['name']
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        contact = request.form['contact']
        email = request.form['email']
        license_no = request.form['license_no']
        joined_date = request.form['joined_date']
        address = request.form['address']
        
        # Validate Username & Password: 15 chars max, alphanumeric letters & numbers only, no special chars
        if len(username) > 15 or not re.match(r'^[a-zA-Z0-9]{1,15}$', username):
            conn.close()
            flash('Username must contain only letters and numbers, maximum 15 characters (no special characters allowed).', 'danger')
            return redirect(url_for('users'))

        if password and password != '••••••••':
            if len(password) > 15 or not re.match(r'^[a-zA-Z0-9]{1,15}$', password):
                conn.close()
                flash('Password must contain only letters and numbers, maximum 15 characters (no special characters allowed).', 'danger')
                return redirect(url_for('users'))
            if int(user_db_id) == session.get('user_id'):
                conn.execute('''
                    UPDATE users 
                    SET name = ?, username = ?, password = ?, role = ?, contact = ?, email = ?, license_no = ?, joined_date = ?, address = ?
                    WHERE id = ?
                ''', (name, username, password, role, contact, email, license_no, joined_date, address, user_db_id))
            else:
                conn.execute('''
                    UPDATE users 
                    SET name = ?, username = ?, role = ?, contact = ?, email = ?, license_no = ?, joined_date = ?, address = ?
                    WHERE id = ?
                ''', (name, username, role, contact, email, license_no, joined_date, address, user_db_id))
            conn.commit()
            
            # Update session details if editing own profile
            if int(user_db_id) == session.get('user_id'):
                session['name'] = name
                session['role'] = role
            flash('User details updated successfully!')
        else:
            # Check maximum user limit (max 5 users)
            current_user_count = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
            if current_user_count >= 5:
                conn.close()
                flash('Maximum user limit reached (5 users max). Cannot create additional accounts.', 'danger')
                return redirect(url_for('users'))

            # Generate next user ID (e.g. USR-003)
            last_user = conn.execute("SELECT user_id FROM users ORDER BY id DESC LIMIT 1").fetchone()
            if last_user:
                try:
                    last_num = int(last_user['user_id'].split('-')[1])
                    next_user_id = f"USR-{last_num + 1:03d}"
                except (IndexError, ValueError):
                    next_user_id = "USR-003"
            else:
                next_user_id = "USR-001"
                
            conn.execute('''
                INSERT INTO users (user_id, name, username, password, role, contact, email, license_no, joined_date, address)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (next_user_id, name, username, password, role, contact, email, license_no, joined_date, address))
            conn.commit()
            flash('New user added successfully!')
            
        conn.close()
        return redirect(url_for('users'))
        
    users_list = conn.execute('SELECT * FROM users').fetchall()
    conn.close()
    return render_template('users.html', users=users_list, active_page='users')

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    conn = get_db_connection()
    
    if request.method == 'POST':
        if session.get('role') == 'Staff':
            keys = [
                'dark_mode', 'font_size', 'high_contrast', 'reduce_motion', 
                'screen_reader', 'show_generic', 'low_stock_alerts', 
                'expiry_alerts', 'po_alerts', 'sales_notifications', 
                'auto_logout', 'require_password'
            ]
        else:
            keys = [
                'pharmacy_name', 'address', 'contact', 'email', 
                'dark_mode', 'font_size', 'high_contrast', 'reduce_motion', 
                'screen_reader', 'show_generic', 'low_stock_alerts', 
                'expiry_alerts', 'po_alerts', 'sales_notifications', 
                'low_stock_threshold', 'near_expiry_warning', 
                'default_currency', 'auto_logout', 'require_password'
            ]
        for key in keys:
            if key in ['dark_mode', 'high_contrast', 'reduce_motion', 'screen_reader', 'show_generic', 
                       'low_stock_alerts', 'expiry_alerts', 'po_alerts', 'sales_notifications', 'require_password']:
                val = 'true' if key in request.form else 'false'
            else:
                val = request.form.get(key, '')
            conn.execute('UPDATE settings SET value = ? WHERE key = ?', (val, key))
            
        conn.commit()
        conn.close()
        flash('Settings saved successfully!')
        return redirect(url_for('settings'))
        
    settings_raw = conn.execute('SELECT * FROM settings').fetchall()
    conn.close()
    
    settings_dict = {r['key']: r['value'] for r in settings_raw}
    return render_template('settings.html', settings=settings_dict, active_page='settings')

@app.route('/notifications/mark-seen', methods=['POST'])
def mark_notifications_seen():
    conn = get_db_connection()
    conn.execute('UPDATE notifications SET seen = 1 WHERE seen = 0')
    conn.commit()
    conn.close()
    return {'status': 'success'}

@app.route('/reports')
def reports():
    conn = get_db_connection()
    categories = ['Medicines', 'Vitamins & Supplements', 'Medical Supplies', 'Personal Care', 'Baby Care', 'Medical Devices', 'First Aid', 'Skin Care']
    
    # 1 & 2. Stock by Category
    cat_stock_raw = conn.execute("SELECT category, SUM(stock) as total_stock FROM inventory GROUP BY category").fetchall()
    cat_stock_map = {r['category']: r['total_stock'] for r in cat_stock_raw}
    cat_stock_values = [cat_stock_map.get(cat, 0) for cat in categories]
    
    # 3. Top 10 Products by Stock
    top_products_raw = conn.execute("SELECT brand, stock FROM inventory ORDER BY stock DESC LIMIT 10").fetchall()
    top_products_labels = [r['brand'] for r in top_products_raw]
    top_products_values = [r['stock'] for r in top_products_raw]
    
    # 4. Low Stock Categories
    low_stock_cat_raw = conn.execute("""
        SELECT category, COUNT(*) as low_count 
        FROM inventory 
        WHERE stock < reorder_point 
        GROUP BY category
    """).fetchall()
    low_stock_cat_map = {r['category']: r['low_count'] for r in low_stock_cat_raw}
    low_stock_cat_values = [low_stock_cat_map.get(cat, 0) for cat in categories]
    
    # 5. Expiring Products by Category
    expiring_cat_raw = conn.execute("""
        SELECT i.category, b.expiry_date
        FROM batches b
        JOIN inventory i ON b.medicine_id = i.id
    """).fetchall()
    
    expiring_cat_map = defaultdict(int)
    for b in expiring_cat_raw:
        try:
            expiry_dt = datetime.strptime(b['expiry_date'], '%m/%d/%Y')
            days_left = (expiry_dt - datetime.now()).days
            if 0 < days_left <= 50:
                expiring_cat_map[b['category']] += 1
        except Exception:
            pass
    expiring_cat_values = [expiring_cat_map.get(cat, 0) for cat in categories]
    
    # 6. Monthly Stock Movement
    movements_raw = conn.execute("""
        SELECT SUBSTR(movement_date, 1, 7) as month, type, SUM(ABS(quantity)) as total_qty
        FROM stock_movements
        GROUP BY month, type
        ORDER BY month ASC
    """).fetchall()
    
    months_set = sorted(list(set(r['month'] for r in movements_raw if r['month'])))
    months_list = months_set[-6:] if len(months_set) > 6 else months_set
    
    stock_in_map = {}
    stock_out_map = {}
    for r in movements_raw:
        m = r['month']
        if r['type'] == 'Stock-In':
            stock_in_map[m] = r['total_qty']
        elif r['type'] == 'Stock-Out':
            stock_out_map[m] = r['total_qty']
            
    stock_in_values = [stock_in_map.get(m, 0) for m in months_list]
    stock_out_values = [stock_out_map.get(m, 0) for m in months_list]
    
    if not months_list:
        months_list = ['No Data']
        stock_in_values = [0]
        stock_out_values = [0]
        
    conn.close()
    
    return render_template('reports.html',
                           active_page='reports',
                           categories=categories,
                           cat_stock_values=cat_stock_values,
                           top_products_labels=top_products_labels,
                           top_products_values=top_products_values,
                           low_stock_cat_values=low_stock_cat_values,
                           expiring_cat_values=expiring_cat_values,
                           months_list=months_list,
                           stock_in_values=stock_in_values,
                           stock_out_values=stock_out_values)

class PDFTemplateReport(FPDF):
    def __init__(self, report_title, period_text=""):
        super().__init__()
        self.report_title = report_title
        self.period_text = period_text

    def header(self):
        # 1. Top Red Header Banner
        self.set_fill_color(192, 57, 43)  # Deep Red #c0392b
        self.rect(0, 0, 210, 24, 'F')

        logo_path = os.path.join(app.static_folder, 'images', 'logo_about_us.png')
        if os.path.exists(logo_path):
            self.image(logo_path, x=75, y=3, w=60)
        else:
            self.set_font('helvetica', 'B', 18)
            self.set_text_color(255, 255, 255)
            self.set_xy(0, 6)
            self.cell(210, 10, 'FARMACIA ni DOK', align='C')

        # 2. Sub-header Dark Ribbon
        self.set_fill_color(17, 17, 17)  # Dark Ribbon
        self.rect(0, 24, 210, 7, 'F')
        self.set_font('helvetica', '', 7.5)
        self.set_text_color(255, 255, 255)
        self.set_xy(0, 24)
        self.cell(210, 7, 'Branch : La Residencia, Pio Cruzcrosa, Calumpit Bulacan | farmacianidok@gmail.com | 0917-000-0000', align='C')

        # Reset text color and positioning for page content
        self.set_text_color(0, 0, 0)
        self.set_y(35)

    def footer(self):
        self.set_y(-15)
        self.set_font('helvetica', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', align='C')


@app.route('/reports/download', methods=['POST'])
def download_report():
    if session.get('role') == 'Staff':
        return redirect(url_for('dashboard'))
    report_type = request.form.get('report_type')
    frequency = request.form.get('frequency')
    period_week = request.form.get('period_week')
    period_month = request.form.get('period_month')
    period_year = request.form.get('period_year')
    file_format = request.form.get('format')
    
    start_dt, end_dt = parse_date_range(frequency, period_week, period_month, period_year)
    
    conn = get_db_connection()
    
    headers = []
    rows = []
    title = ""
    
    if report_type == 'inventory':
        title = "Inventory Stock Report"
        headers = ["ID", "Generic Name", "Brand Name", "Category", "Stock", "Reorder Point", "Status"]
        items = conn.execute("SELECT * FROM inventory WHERE archived_at IS NULL AND deleted_at IS NULL").fetchall()
        for item in items:
            rows.append([item['id'], item['generic'], item['brand'], item['category'], item['stock'], item['reorder_point'], item['status']])
            
    elif report_type == 'sales':
        title = "Sales Report"
        headers = ["Sale ID", "Medicine", "Qty Sold", "Sold By", "Date"]
        date_cond, date_params = build_date_filter("s.sale_date")
        sales_data = conn.execute(f"""
            SELECT s.id, i.brand as medicine, s.qty, s.sold_by, s.sale_date 
            FROM sales s 
            JOIN inventory i ON s.medicine_id = i.id
            WHERE 1=1 {date_cond}
            ORDER BY s.id DESC
        """, date_params).fetchall()
        for item in sales_data:
            rows.append([item['id'], item['medicine'], item['qty'], item['sold_by'], item['sale_date']])
            
    elif report_type == 'stock_movements':
        title = "Stock Movements Report"
        headers = ["Movement ID", "Type", "Medicine", "Batch", "Qty", "Date", "Reference"]
        date_cond, date_params = build_date_filter("sm.movement_date")
        movements = conn.execute(f"""
            SELECT sm.id, sm.type, i.brand as medicine, sm.batch_id, sm.quantity, sm.movement_date, sm.reference
            FROM stock_movements sm
            JOIN inventory i ON sm.medicine_id = i.id
            WHERE 1=1 {date_cond}
            ORDER BY sm.id DESC
        """, date_params).fetchall()
        for item in movements:
            rows.append([item['id'], item['type'], item['medicine'], item['batch_id'], item['quantity'], item['movement_date'], item['reference']])
            
    elif report_type == 'expiry_monitoring':
        title = "Expiry Monitoring Report"
        headers = ["Batch ID", "Medicine", "Expiry Date", "Current Qty", "Status"]
        batches = conn.execute("""
            SELECT b.id, i.brand as medicine, b.expiry_date, b.current_qty, b.status
            FROM batches b
            JOIN inventory i ON b.medicine_id = i.id
            ORDER BY b.id DESC
        """).fetchall()
        for item in batches:
            rows.append([item['id'], item['medicine'], item['expiry_date'], item['current_qty'], item['status']])
            
    elif report_type == 'suppliers':
        title = "Suppliers Report"
        headers = ["ID", "Supplier Name", "Contact Person", "Phone", "Email", "Address"]
        suppliers = conn.execute("SELECT * FROM suppliers").fetchall()
        for item in suppliers:
            rows.append([item['id'], item['name'], item['contact_person'], item['phone'], item['email'], item['address']])
            
    conn.close()
    filename = f"{report_type}_report"

    download_time_str = datetime.now().strftime('%B %d, %Y at %I:%M %p')

    if file_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        ws.append([title])
        ws.append([f"Downloaded on: {download_time_str}"])
        if start_dt and end_dt:
            ws.append([f"Period: {start_dt.strftime('%m/%d/%Y')} - {end_dt.strftime('%m/%d/%Y')}"])
        ws.append([])
        
        ws.append(headers)
        for r in rows:
            ws.append(r)
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{filename}.xlsx"
        )
        
    elif file_format == 'pdf':
        period_text = f"Period: {start_dt.strftime('%m/%d/%Y')} - {end_dt.strftime('%m/%d/%Y')}" if (start_dt and end_dt) else ""
        pdf = PDFTemplateReport(title, period_text)
        pdf.alias_nb_pages()
        pdf.add_page()
        
        pdf.set_font("helvetica", "B", 13)
        pdf.cell(0, 8, title.upper(), align="C")
        pdf.ln(6)
        
        pdf.set_font("helvetica", "I", 8.5)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(0, 5, f"Downloaded on: {download_time_str}", align="C")
        pdf.ln(5)
        
        if period_text:
            pdf.cell(0, 5, period_text, align="C")
            pdf.ln(5)
            
        pdf.set_text_color(0, 0, 0)
        pdf.ln(3)

        # Header Row
        pdf.set_fill_color(192, 57, 43)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("helvetica", "B", 9)
        col_width = 190 / len(headers)
        
        for h in headers:
            pdf.cell(col_width, 8, str(h), border=1, align="C", fill=True)
        pdf.ln()
        
        # Rows with alternating colors
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("helvetica", size=8.5)
        fill = False
        for row in rows:
            if pdf.get_y() > 260:
                pdf.add_page()
                pdf.set_fill_color(192, 57, 43)
                pdf.set_text_color(255, 255, 255)
                pdf.set_font("helvetica", "B", 9)
                for h in headers:
                    pdf.cell(col_width, 8, str(h), border=1, align="C", fill=True)
                pdf.ln()
                pdf.set_text_color(0, 0, 0)
                pdf.set_font("helvetica", size=8.5)
            
            if fill:
                pdf.set_fill_color(248, 250, 252)
            else:
                pdf.set_fill_color(255, 255, 255)
                
            for val in row:
                pdf.cell(col_width, 7.5, str(val), border=1, align="C", fill=True)
            pdf.ln()
            fill = not fill
            
        pdf_data = pdf.output()
        output = BytesIO(pdf_data)
        
        return send_file(
            output,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"{filename}.pdf"
        )
        
    return "Invalid format", 400

@app.route('/backup_db')
def backup_db():
    if 'user_id' not in session or session.get('role') == 'Staff':
        return jsonify({'success': False, 'error': 'Permission denied. Admin privileges required.'}), 403
    
    from database import get_db_path
    import tempfile, sqlite3, base64
    db_path = get_db_path()
    
    if not os.path.exists(db_path):
        return jsonify({'success': False, 'error': 'Database file not found.'}), 404
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_filename = f"farmacia_backup_{timestamp}.db"
    
    temp_dir = tempfile.gettempdir()
    temp_backup_path = os.path.join(temp_dir, backup_filename)
    
    try:
        src_conn = sqlite3.connect(db_path)
        dst_conn = sqlite3.connect(temp_backup_path)
        src_conn.backup(dst_conn)
        dst_conn.close()
        src_conn.close()
        
        with open(temp_backup_path, 'rb') as f:
            db_bytes = f.read()
        base64_data = base64.b64encode(db_bytes).decode('utf-8')
        
        user_downloads = os.path.join(os.path.expanduser('~'), 'Downloads')
        downloads_saved_path = None
        if os.path.exists(user_downloads):
            downloads_saved_path = os.path.join(user_downloads, backup_filename)
            with open(downloads_saved_path, 'wb') as f_dl:
                f_dl.write(db_bytes)
        
        if os.path.exists(temp_backup_path):
            try:
                os.remove(temp_backup_path)
            except Exception:
                pass
                
        return jsonify({
            'success': True,
            'filename': backup_filename,
            'base64': base64_data,
            'downloads_path': downloads_saved_path or backup_filename
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/restore_db', methods=['POST'])
def restore_db():
    if 'user_id' not in session or session.get('role') == 'Staff':
        flash('Permission denied. Admin privileges required to restore database.')
        return redirect(url_for('settings'))
    
    if 'backup_file' not in request.files:
        flash('No backup file selected.')
        return redirect(url_for('settings'))
    
    file = request.files['backup_file']
    if file.filename == '':
        flash('No backup file selected.')
        return redirect(url_for('settings'))
    
    if not (file.filename.endswith('.db') or file.filename.endswith('.sqlite') or file.filename.endswith('.bak')):
        flash('Invalid backup file format. Please upload a valid .db or .sqlite backup file.')
        return redirect(url_for('settings'))
    
    from database import get_db_path
    import tempfile, sqlite3, time
    db_path = get_db_path()
    
    temp_dir = tempfile.gettempdir()
    temp_file_path = os.path.join(temp_dir, f"temp_restore_{int(time.time())}.db")
    
    try:
        file.save(temp_file_path)
        
        source_conn = sqlite3.connect(temp_file_path)
        check_res = source_conn.execute("PRAGMA quick_check").fetchone()
        
        if check_res and check_res[0] == 'ok':
            dest_conn = sqlite3.connect(db_path)
            source_conn.backup(dest_conn)
            dest_conn.close()
            source_conn.close()
            
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
                
            flash('Database successfully restored from backup!')
        else:
            source_conn.close()
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
            flash('Corrupted or invalid database backup file. Restore cancelled.')
            
    except Exception as e:
        if os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except Exception:
                pass
        flash(f'Database restore failed: {str(e)}')
        
    return redirect(url_for('settings'))

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
